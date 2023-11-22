import pandas as pd
import numpy as np
from datetime import *
from openpyxl import *
from openpyxl.styles import Alignment
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score

# 函数化你的代码
def process_real_time(real_time, machine_time, Rtime, usecols, sheet, start_col,gas_column):
    date1 = datetime.strptime(real_time, "%Y/%m/%d %H:%M:%S")
    date2 = datetime.strptime(machine_time, "%Y/%m/%d %H:%M:%S")

    duration = date1 - date2

    Mtime = []
    for i in range(len(Rtime)):
        tmp = (datetime.strptime(Rtime[i], "%Y/%m/%d %H:%M:%S")) - (duration)
        Mtime.append(str(tmp))

    data = pd.DataFrame({'機器時間': Mtime, '實際時間': Rtime})

    df = pd.read_csv(file_dir, sep='\s+', usecols=usecols)
    df_new = df.dropna()

    df_new['TIME'] = pd.to_datetime(df_new['TIME'], format='%H:%M:%S', errors='coerce').dt.time
    df_new['DATE'] = pd.to_datetime(df_new['DATE'], format='%Y-%m-%d', errors='coerce')

    df_new['DATETIME'] = pd.to_datetime(df_new['DATE'].astype(str) + ' ' + df_new['TIME'].astype(str),
                                         errors='coerce')

    for idx, current_time in enumerate(Rtime):
        start_time = pd.to_datetime(Mtime[idx])
        start_time = start_time + timedelta(seconds=59)
        end_time = start_time + timedelta(minutes=3, seconds=1)

        selected_data = df_new[(df_new["DATETIME"] > start_time) & (df_new["DATETIME"] < end_time)]

        regression_results = []
        max_r2 = float('-inf')
        max_r2_record = []

        for i in range(len(selected_data) - 120):
            window_data = selected_data[gas_column].iloc[i:i + 120].values.reshape(-1, 1)
            window_time = np.arange(120).reshape(-1, 1)

            model = LinearRegression()
            model.fit(window_time, window_data)

            a, b = model.coef_[0][0], model.intercept_[0]

            predicted_values = model.predict(window_time)
            r2 = r2_score(window_data, predicted_values)

            if r2 > max_r2:
                max_r2 = r2
                max_r2_record = [{'start_time': selected_data['TIME'].iloc[i], 'a': a, 'r2': r2}]

            regression_results.append({'start_time': selected_data['TIME'].iloc[i], 'a': a, 'b': b, 'r2': r2})
        
        max_r2_record = max_r2_record[0]  # 取出 dict 資訊
        max_r2_start_time = max_r2_record['start_time']
        max_r2_a = max_r2_record['a']
        max_r2_r2 = max_r2_record['r2']
        # 將 max_r2_start_time 轉換為時間
        max_r2_start_time = datetime.combine(date.today(), max_r2_start_time)  # 假設日期是當天的日期
        print(max_r2_record)
        # 創建包含 max_r2_start_time 以及之後 119 秒的時間的列表
        time_series = [max_r2_start_time + timedelta(seconds=i) for i in range(120)]

        # 寫入 Start Time
        sheet.cell(row=1, column=start_col + idx, value=f"Start Time {idx + 1}")

        # 寫入時間序列
        for i, time in enumerate(time_series, start=2):
            sheet.cell(row=i, column=start_col + idx, value=time.time())

        # 寫入 a 和 R^2
        sheet.cell(row=122, column=start_col + idx, value=f"a {idx + 1}")
        sheet.cell(row=123, column=start_col + idx, value=max_r2_a)
        sheet.cell(row=124, column=start_col + idx, value=f"R^2 {idx + 1}")
        sheet.cell(row=125, column=start_col + idx, value=max_r2_r2)

    # 設定儲存格格式
    for row in sheet.iter_rows(min_row=1, max_row=125, min_col=start_col, max_col=start_col + len(Rtime)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

# 人工輸入區
machine_num = "1450"  # 機器代號
file_dir = "TG10-01450-2023-10-23T194200.data.txt"  # 你要用的 data
usecols = ['DATE', 'TIME', 'CO2']  # 你要抓的資料(記得要改格式)
gas_column=['CO2']
Rtime = ["2023/10/25 15:54:50",
         "2023/10/25 16:02:50",
         "2023/10/25 16:14:10",
         "2023/10/25 16:24:30",
         "2023/10/25 16:34:00",
         "2023/10/25 16:44:50"]  # 實際時間

# 創建 Workbook
wb = Workbook()
ws = wb.active

# 呼叫函数处理每个实际时间
process_real_time("2023/10/25 20:51:10", "2023/10/24 21:10:29", Rtime, usecols, ws, 1,gas_column)

# 儲存 Excel
output_excel = "output_result_openpyxl.xlsx"
wb.save(output_excel)
print("輸出成功")
