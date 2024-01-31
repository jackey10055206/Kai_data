import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from openpyxl.styles import Alignment
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from openpyxl import Workbook

from process_data import machine_num, machine_series, file_dir, usecols, gas_column, gas_type, Real_time, Machine_time, Rtime

start_col = 1

# 函数化你的代码
def process_real_time():
    date1 = datetime.strptime(Real_time, "%Y/%m/%d %H:%M:%S")
    date2 = datetime.strptime(Machine_time, "%Y/%m/%d %H:%M:%S")

    duration = date1 - date2

    Mtime = []
    for i in range(len(Rtime)):
        tmp = (datetime.strptime(Rtime[i], "%Y/%m/%d %H:%M:%S")) - (duration)
        Mtime.append(str(tmp))

    data = pd.DataFrame({'機器時間': Mtime, '實際時間': Rtime})

    df = pd.read_csv(file_dir, sep='\s+', usecols=usecols)
    df_new = df.dropna()

    df_new['TIME'] = pd.to_datetime(df_new['TIME'], format='%H:%M:%S', errors='coerce').dt.time
    df_new['DATE'] = pd.to_datetime(df_new['DATE'], format='%Y-%m-%d', errors='coerce')
    df_new['DATETIME'] = pd.to_datetime(df_new['DATE'].astype(str) + ' ' + df_new['TIME'].astype(str),
                                         errors='coerce')

    for idx, current_time in enumerate(Rtime):
        start_time = pd.to_datetime(Mtime[idx])
        start_time = start_time + timedelta(seconds=59)
        end_time = start_time + timedelta(minutes=3, seconds=1)
       
        selected_data = df_new[(df_new["DATETIME"] > start_time) & (df_new["DATETIME"] < end_time)]
    
        regression_results = []
        max_r2 = float('-inf')
        max_r2_record = []

        for i in range(len(selected_data) - 120):
            window_data = selected_data[gas_column].iloc[i:i + 120].values.reshape(-1, 1)
            window_time = np.arange(120).reshape(-1, 1)

            model = LinearRegression()
            model.fit(window_time, window_data)

            a, b = model.coef_[0][0], model.intercept_[0]

            predicted_values = model.predict(window_time)
            r2 = r2_score(window_data, predicted_values)
            
            if r2 > max_r2:
                max_r2 = r2
                max_r2_record = [{'start_time': selected_data['TIME'].iloc[i], 'a': a, 'r2': r2, 'window_data': window_data}]
                
            regression_results.append({'start_time': selected_data['TIME'].iloc[i], 'a': a, 'b': b, 'r2': r2})
        
        
        if not max_r2_record:
            print(f"No record for time {current_time}")
        else:           
            max_r2_record = max_r2_record[0]  # 取出 dict 資訊
            max_r2_start_time = max_r2_record['start_time']
            max_r2_a = max_r2_record['a']
            max_r2_r2 = max_r2_record['r2']
            # 將 max_r2_start_time 轉換為時間
            max_r2_start_time = datetime.combine(date.today(), max_r2_start_time)  # 假設日期是當天的日期
            # 創建包含 max_r2_start_time 以及之後 119 秒的時間的列表
            time_series = [max_r2_start_time + timedelta(seconds=i) for i in range(120)]

            # 寫入 Start Time
            sheet.cell(row=1, column=start_col + idx * 2, value=f"Start Time {idx + 1}")

            # 寫入時間序列和濃度值
            for i, (time, concentration) in enumerate(zip(time_series, max_r2_record['window_data'])):
                sheet.cell(row=i + 2, column=start_col + idx * 2, value=time.time())
                sheet.cell(row=i + 2, column=start_col + idx * 2 + 1, value=float(concentration[0]))

            # 寫入 a 和 R^2
            sheet.cell(row=122, column=start_col + idx * 2, value=f"a {idx + 1}")
            sheet.cell(row=123, column=start_col + idx * 2, value=max_r2_a)
            sheet.cell(row=124, column=start_col + idx * 2, value=f"R^2 {idx + 1}")
            sheet.cell(row=125, column=start_col + idx * 2, value=max_r2_r2)

            # 設定儲存格格式
            for row in sheet.iter_rows(min_row=1, max_row=125, min_col=start_col, max_col=start_col + len(Rtime) * 2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

# 創建 Workbook
wb = Workbook()
sheet = wb.active

# 呼叫函数处理每个实际时间
process_real_time()

# 儲存 Excel
output_excel = f"{machine_num}_{machine_series}_{Rtime[0][5:7]}_{Rtime[0][8:10]}_{gas_column[0]}_{gas_type}.xlsx"
wb.save(output_excel)
print("輸出成功")
